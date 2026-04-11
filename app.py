import streamlit as st
from transformers import pipeline
from PIL import Image
from gtts import gTTS
import tempfile
import os
import requests
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import random
import io
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -------------------------
# PAGE CONFIG
# -------------------------
st.set_page_config(
    page_title="AI Food Health Analyzer",
    layout="wide",
    page_icon="🥗"
)

# -------------------------
# CUSTOM CSS
# -------------------------
st.markdown("""
<style>
    /* Main background */
    .stApp { background-color: #0f1117; }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a1f2e 0%, #151b2b 100%);
        border-right: 1px solid #2a3045;
    }

    /* Cards */
    .metric-card {
        background: linear-gradient(135deg, #1e2535 0%, #252d42 100%);
        border: 1px solid #2e3a55;
        border-radius: 12px;
        padding: 18px 20px;
        margin: 6px 0;
        text-align: center;
    }
    .metric-card .label {
        font-size: 12px;
        color: #7b8cad;
        text-transform: uppercase;
        letter-spacing: 1px;
        margin-bottom: 4px;
    }
    .metric-card .value {
        font-size: 28px;
        font-weight: 700;
        color: #e2e8f0;
    }
    .metric-card .unit {
        font-size: 13px;
        color: #7b8cad;
    }

    /* Section headers */
    .section-header {
        font-size: 20px;
        font-weight: 700;
        color: #e2e8f0;
        padding: 10px 0 6px 0;
        border-bottom: 2px solid #3b82f6;
        margin-bottom: 16px;
        display: inline-block;
    }

    /* Food tag */
    .food-tag {
        display: inline-block;
        background: linear-gradient(90deg, #1d4ed8, #2563eb);
        color: white;
        border-radius: 20px;
        padding: 5px 14px;
        margin: 4px;
        font-size: 14px;
        font-weight: 600;
    }

    /* Meal card */
    .meal-card {
        background: linear-gradient(135deg, #1a2744 0%, #1e2f52 100%);
        border: 1px solid #2e4175;
        border-left: 4px solid #3b82f6;
        border-radius: 10px;
        padding: 16px 18px;
        margin: 10px 0;
    }
    .meal-card h4 {
        color: #93c5fd;
        margin: 0 0 8px 0;
        font-size: 16px;
    }
    .meal-card p {
        color: #cbd5e1;
        margin: 0;
        font-size: 14px;
        line-height: 1.6;
    }

    /* Status badges */
    .badge-green  { background:#065f46; color:#6ee7b7; padding:3px 10px; border-radius:20px; font-size:13px; font-weight:600; }
    .badge-yellow { background:#78350f; color:#fcd34d; padding:3px 10px; border-radius:20px; font-size:13px; font-weight:600; }
    .badge-red    { background:#7f1d1d; color:#fca5a5; padding:3px 10px; border-radius:20px; font-size:13px; font-weight:600; }

    /* Tip box */
    .tip-box {
        background: #1e2d45;
        border-left: 4px solid #f59e0b;
        border-radius: 8px;
        padding: 12px 16px;
        margin: 6px 0;
        color: #fde68a;
        font-size: 14px;
    }

    /* Per-food nutrient row */
    .nutrient-row {
        background: #1c2333;
        border: 1px solid #2a3550;
        border-radius: 8px;
        padding: 10px 16px;
        margin: 5px 0;
        display: flex;
        justify-content: space-between;
        align-items: center;
    }

    /* Hide default metric delta */
    [data-testid="stMetricDelta"] { display: none; }

    /* Upload area */
    [data-testid="stFileUploader"] {
        background: #1a1f2e;
        border: 2px dashed #2e3a55;
        border-radius: 12px;
        padding: 10px;
    }

    /* Divider */
    hr { border-color: #2a3045; }
</style>
""", unsafe_allow_html=True)

# -------------------------
# HEADER
# -------------------------
col_logo, col_title = st.columns([1, 11])
with col_title:
    st.markdown("""
    <div style='padding:10px 0'>
        <span style='font-size:36px; font-weight:800; color:#e2e8f0;'>🥗 GenAI Food & Health Analyzer</span><br>
        <span style='color:#7b8cad; font-size:15px;'>Upload food images → Get instant nutrition analysis, BMI insights & personalized meal recommendations</span>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<hr>", unsafe_allow_html=True)

# -------------------------
# SIDEBAR
# -------------------------
st.sidebar.markdown("""
<div style='text-align:center; padding:10px 0 20px 0;'>
    <span style='font-size:22px; font-weight:700; color:#e2e8f0;'>👤 Your Profile</span>
</div>
""", unsafe_allow_html=True)

age    = st.sidebar.number_input("Age", 5, 80, value=22)
weight = st.sidebar.number_input("Weight (kg)", 20, 150, value=65)
height_ft = st.sidebar.number_input("Height (feet)", 1.0, 7.0, value=5.7, step=0.1)
height = round(height_ft * 30.48)
gender = st.sidebar.selectbox("Gender", ["Male", "Female"])
goal   = st.sidebar.selectbox("Health Goal", ["Weight Loss", "Muscle Gain", "Maintain"])
activity = st.sidebar.selectbox("Activity Level", ["Sedentary", "Lightly Active", "Moderately Active", "Very Active"])

activity_multiplier = {"Sedentary": 1.2, "Lightly Active": 1.375,
                       "Moderately Active": 1.55, "Very Active": 1.725}

# Sidebar BMI preview
h_m = height / 100
bmi_preview = round(weight / (h_m * h_m), 1)
if   bmi_preview < 18.5: bmi_color, bmi_label = "#fcd34d", "Underweight"
elif bmi_preview < 25:   bmi_color, bmi_label = "#6ee7b7", "Normal"
elif bmi_preview < 30:   bmi_color, bmi_label = "#fcd34d", "Overweight"
else:                    bmi_color, bmi_label = "#fca5a5", "Obese"

st.sidebar.markdown(f"""
<div style='background:#1c2333; border:1px solid #2a3550; border-radius:10px; padding:14px; margin-top:16px; text-align:center;'>
    <div style='font-size:12px; color:#7b8cad; margin-bottom:4px;'>YOUR BMI</div>
    <div style='font-size:36px; font-weight:800; color:{bmi_color};'>{bmi_preview}</div>
    <div style='font-size:13px; color:{bmi_color};'>{bmi_label}</div>
</div>
""", unsafe_allow_html=True)

# -------------------------
# LOAD MODEL
# -------------------------
@st.cache_resource
def load_model():
    return pipeline("image-classification", model="nateraw/food")

classifier = load_model()

# -------------------------
# API KEY
# -------------------------
USDA_API_KEY = st.secrets["USDA_API_KEY"]

# -------------------------
# CALCULATIONS
# -------------------------
def calculate_bmi(weight, height):
    h = height / 100
    bmi = weight / (h * h)
    if   bmi < 18.5: return round(bmi, 2), "Underweight"
    elif bmi < 25:   return round(bmi, 2), "Normal"
    elif bmi < 30:   return round(bmi, 2), "Overweight"
    else:            return round(bmi, 2), "Obese"

def calculate_bmr(weight, height, age, gender):
    if gender == "Male": return 10*weight + 6.25*height - 5*age + 5
    else:                return 10*weight + 6.25*height - 5*age - 161

def get_rda(age, weight, height, gender, activity_level="Moderately Active"):
    bmr      = calculate_bmr(weight, height, age, gender)
    mult     = activity_multiplier.get(activity_level, 1.55)
    calories = round(bmr * mult)
    protein  = round(0.8 * weight, 1)
    carbs    = round((0.5 * calories) / 4, 1)
    fat      = round((0.25 * calories) / 9, 1)
    return {"calories": calories, "protein": protein, "carbs": carbs, "fat": fat}

def clean_food(food):
    mapping = {"cheeseburger": "burger", "french fries": "fries", "bread pudding": "bread"}
    return mapping.get(food.lower(), food)

def get_nutrition(food):
    try:
        food = clean_food(food)
        url  = f"https://api.nal.usda.gov/fdc/v1/foods/search?query={food}&api_key={USDA_API_KEY}"
        res  = requests.get(url).json()
        nutrients = {"calories": 0, "protein": 0, "fat": 0, "carbs": 0, "vitamins": 0}
        foods = res.get("foods", [])
        if not foods: return nutrients
        for item in foods[0]["foodNutrients"]:
            name = item["nutrientName"].lower()
            if   "energy"       in name: nutrients["calories"] = item["value"]
            elif "protein"      in name: nutrients["protein"]  = item["value"]
            elif "fat"          in name: nutrients["fat"]      = item["value"]
            elif "carbohydrate" in name: nutrients["carbs"]    = item["value"]
            elif "vitamin"      in name: nutrients["vitamins"] += item["value"]
        return nutrients
    except:
        return {"calories": 0, "protein": 0, "fat": 0, "carbs": 0, "vitamins": 0}

def calculate_total(foods):
    total   = {"calories": 0, "protein": 0, "fat": 0, "carbs": 0, "vitamins": 0}
    details = []
    for food in foods:
        data = get_nutrition(food)
        details.append((food, data))
        for k in total: total[k] += data[k]
    return total, details

def analyze_deficiency(total, rda):
    tips = []
    if total["protein"] < 0.8 * rda["protein"]:  tips.append(("⚠️ Protein low",  "Eat eggs, dal, paneer, chicken."))
    if total["carbs"]   < 0.8 * rda["carbs"]:    tips.append(("⚠️ Carbs low",    "Add rice, oats, sweet potato."))
    if total["fat"]     < 0.8 * rda["fat"]:      tips.append(("⚠️ Healthy fats low", "Add nuts, avocado, olive oil."))
    if total["calories"] > 1.2 * rda["calories"]: tips.append(("🔴 Calorie excess", "Reduce fried/junk food intake."))
    if total["calories"] < 0.8 * rda["calories"]: tips.append(("🔴 Calorie deficit", "Increase meal portions."))
    return tips if tips else [("✅ Balanced", "Your diet looks nutritionally balanced!")]

def evaluate_food_health(foods, total, bmi, goal, age):
    unhealthy  = ["burger", "fries", "pizza", "donut", "hot dog"]
    feedback   = []
    junk_items = [f for f in foods if any(j in f.lower() for j in unhealthy)]

    if   bmi > 25:   feedback.append("You are overweight. High-calorie foods increase fat storage and may lead to heart disease.")
    elif bmi < 18.5: feedback.append("You are underweight. Low-nutrient foods will not help in healthy weight gain.")
    else:            feedback.append("Your BMI is normal. Maintaining diet quality is important.")

    if junk_items:
        feedback.append(f"Detected unhealthy items: {', '.join(junk_items)}.")
        feedback.append("These contain high unhealthy fats, sugar, and low nutrients.")
        if   age < 18:  feedback.append("May affect growth and immunity in your age group.")
        elif age <= 40: feedback.append("May lead to weight gain and poor metabolism.")
        else:           feedback.append("Increases heart and cholesterol risks significantly.")
        if goal == "Weight Loss":  feedback.append("Not suitable for your weight loss goal.")
        if goal == "Muscle Gain":  feedback.append("Does not support muscle growth.")
        feedback.append("Replace with vegetables, fruits, whole grains, and lean proteins.")
    else:
        feedback.append("Foods chosen are healthy in moderation. Good choices!")
    return feedback

def evaluate_goal(bmi, goal):
    if bmi < 18.5:
        if goal == "Weight Loss": return "❌ You are underweight. Weight loss is not recommended.", "red"
        else: return "✅ Your goal is appropriate for your current BMI.", "green"
    elif bmi > 25:
        if goal == "Muscle Gain": return "⚠️ You are overweight. Consider fat loss before muscle gain.", "yellow"
        elif goal == "Maintain":  return "⚠️ Consider switching to Weight Loss goal.", "yellow"
        else: return "✅ Your goal aligns with your health needs.", "green"
    else:
        return "✅ Your goal is perfectly suited to your current health.", "green"

# -------------------------
# MEAL RECOMMENDATIONS (dynamic — scaled to person's calorie target)
# -------------------------

# Fractions define what % of daily calories each meal slot should cover
MEAL_SLOT_FRACTIONS = {
    "Weight Loss": {"Breakfast": 0.25, "Lunch": 0.35, "Dinner": 0.28, "Snacks": 0.12},
    "Muscle Gain": {"Breakfast": 0.28, "Lunch": 0.35, "Dinner": 0.28, "Snacks": 0.09},
    "Maintain":    {"Breakfast": 0.25, "Lunch": 0.35, "Dinner": 0.30, "Snacks": 0.10},
}

# Each option has its own calorie multiplier (relative to slot target) and a quantity-scaling fn.
# Multipliers:  1.0 = on-target,  0.9 = lighter alternative,  0.8 = lightest option
# desc_fn receives the option's actual kcal so quantities differ between cards.
MEAL_TEMPLATES = {
    "Weight Loss": {
        "Breakfast": [
            (1.00, "Veggie Egg White Omelette",  lambda c: f"{max(3,round(c/55))} egg whites, spinach, tomato, mushrooms, 1 slice wholegrain toast"),
            (0.90, "Greek Yogurt Parfait",        lambda c: f"{max(150,round(c/0.9))}g low-fat Greek yogurt, {max(30,round(c/9))}g granola, mixed berries"),
            (0.80, "Overnight Oats",              lambda c: f"{max(40,round(c/4.8))}g rolled oats, 150ml almond milk, chia seeds, sliced strawberries"),
        ],
        "Lunch": [
            (1.00, "Grilled Chicken Salad",       lambda c: f"{max(100,round(c/1.8))}g grilled chicken breast, mixed greens, cherry tomatoes, cucumber, balsamic vinaigrette"),
            (0.90, "Turkey & Avocado Wrap",       lambda c: f"{max(80,round(c/2))}g turkey slices, ½ avocado, lettuce, tomato in 1 wholegrain wrap"),
            (0.85, "Tuna Nicoise Salad",          lambda c: f"{max(90,round(c/1.9))}g canned tuna, boiled egg, green beans, olives, mixed greens, lemon dressing"),
        ],
        "Dinner": [
            (1.00, "Baked Salmon + Steamed Veg",  lambda c: f"{max(120,round(c/1.7))}g baked salmon, broccoli, asparagus, lemon-herb seasoning"),
            (0.88, "Grilled Turkey Breast + Salad",lambda c: f"{max(110,round(c/1.9))}g grilled turkey, large garden salad, olive oil dressing"),
            (0.80, "Zucchini Noodles + Marinara", lambda c: f"{max(200,round(c/1.5))}g zucchini noodles, homemade tomato marinara, {max(50,round(c/5))}g lean ground turkey"),
        ],
        "Snacks": [
            (1.00, "Apple + Almond Butter",       lambda c: f"1 medium apple, {max(10,round(c/19))}g almond butter"),
            (0.85, "Cottage Cheese + Cucumber",   lambda c: f"{max(80,round(c/1.3))}g low-fat cottage cheese, sliced cucumber, black pepper"),
            (0.70, "Celery Sticks + Hummus",      lambda c: f"celery sticks, {max(30,round(c/2.3))}g hummus"),
        ],
    },
    "Muscle Gain": {
        "Breakfast": [
            (1.00, "Scrambled Eggs + Toast",      lambda c: f"{max(3,round(c/75))} whole eggs, {max(2,round(c/175))} slices wholegrain toast, 1 tbsp butter, side of sliced avocado"),
            (0.92, "Protein Pancakes",            lambda c: f"{max(60,round(c/5))}g oat flour, {max(2,round(c/230))} eggs, 1 scoop whey protein, topped with banana & honey"),
            (0.85, "Bagel + Smoked Salmon",       lambda c: f"1 whole wheat bagel, {max(80,round(c/3.5))}g smoked salmon, cream cheese, capers, red onion"),
        ],
        "Lunch": [
            (1.00, "Chicken & Brown Rice Bowl",   lambda c: f"{max(150,round(c/2.3))}g grilled chicken breast, {max(80,round(c/5.2))}g brown rice, roasted veggies, olive oil"),
            (0.92, "Beef Stir Fry + Noodles",     lambda c: f"{max(130,round(c/2.5))}g lean beef strips, {max(75,round(c/5))}g egg noodles, bell peppers, soy sauce, sesame oil"),
            (0.88, "Tuna Pasta Salad",            lambda c: f"{max(70,round(c/4.8))}g wholegrain pasta, {max(100,round(c/2.5))}g tuna, sweetcorn, mayo, mixed greens"),
        ],
        "Dinner": [
            (1.00, "Sirloin Steak + Sweet Potato",lambda c: f"{max(150,round(c/2.0))}g lean sirloin, 1 medium sweet potato, steamed green beans, garlic butter"),
            (0.92, "Baked Chicken Thighs + Quinoa",lambda c: f"{max(140,round(c/2.2))}g baked chicken thighs, {max(70,round(c/5))}g quinoa, roasted broccoli, olive oil"),
            (0.88, "Salmon + Mashed Potato",      lambda c: f"{max(130,round(c/2.4))}g baked salmon, {max(150,round(c/2.2))}g mashed potato (with milk), steamed spinach"),
        ],
        "Snacks": [
            (1.00, "Protein Shake + Banana",      lambda c: f"1 scoop whey protein (shaken with 300ml milk), 1 large banana"),
            (0.85, "Peanut Butter Rice Cakes",    lambda c: f"{max(2,round(c/120))} rice cakes, {max(15,round(c/16))}g peanut butter, honey drizzle"),
            (0.75, "Hard Boiled Eggs + Cheese",   lambda c: f"{max(2,round(c/80))} boiled eggs, {max(20,round(c/22))}g cheddar cheese"),
        ],
    },
    "Maintain": {
        "Breakfast": [
            (1.00, "Avocado Toast + Poached Eggs",lambda c: f"{max(2,round(c/190))} slices sourdough, ½ avocado, {max(1,round(c/370))} poached egg, chilli flakes"),
            (0.92, "Smoothie Bowl",               lambda c: f"{max(150,round(c/1.5))}g blended frozen berries + banana, topped with {max(25,round(c/14))}g granola, chia seeds, sliced kiwi"),
            (0.85, "Wholegrain Cereal + Milk",    lambda c: f"{max(50,round(c/5.5))}g wholegrain cereal, 200ml semi-skimmed milk, handful blueberries"),
        ],
        "Lunch": [
            (1.00, "Grilled Chicken Sandwich",    lambda c: f"{max(100,round(c/2.2))}g grilled chicken, wholegrain baguette, lettuce, tomato, mustard mayo, side salad"),
            (0.92, "Lentil & Vegetable Soup",     lambda c: f"{max(150,round(c/1.8))}g lentil soup, 1 wholegrain roll, mixed salad"),
            (0.88, "Caesar Salad + Grilled Shrimp",lambda c: f"{max(100,round(c/1.9))}g grilled shrimp, romaine lettuce, {max(15,round(c/25))}g parmesan, croutons, Caesar dressing"),
        ],
        "Dinner": [
            (1.00, "Roast Chicken + Roast Veg",   lambda c: f"{max(130,round(c/2.1))}g roast chicken breast, roasted potato, carrots, parsnips, gravy"),
            (0.90, "Pasta Primavera",             lambda c: f"{max(70,round(c/4.5))}g wholegrain pasta, mixed seasonal veggies, olive oil, garlic, parmesan"),
            (0.83, "Baked Cod + New Potatoes",    lambda c: f"{max(120,round(c/2.0))}g baked cod, {max(120,round(c/2.5))}g new potatoes, green beans, lemon-butter sauce"),
        ],
        "Snacks": [
            (1.00, "Mixed Nuts & Dried Fruit",    lambda c: f"{max(20,round(c/3.2))}g mixed nuts, {max(15,round(c/4))}g dried cranberries or raisins"),
            (0.88, "Cheese & Wholegrain Crackers",lambda c: f"{max(20,round(c/3.5))}g cheddar, {max(3,round(c/60))} wholegrain crackers"),
            (0.75, "Fruit & Yogurt Cup",          lambda c: f"{max(100,round(c/1.1))}g low-fat yogurt, sliced apple or grapes"),
        ],
    },
}

def get_meal_recommendations(goal, bmi, age, weight, rda):
    target_cal = rda["calories"]
    fractions  = MEAL_SLOT_FRACTIONS.get(goal, MEAL_SLOT_FRACTIONS["Maintain"])
    templates  = MEAL_TEMPLATES.get(goal, MEAL_TEMPLATES["Maintain"])

    plan = {}
    for slot, frac in fractions.items():
        slot_cal = target_cal * frac
        meals = []
        for mult, name, desc_fn in templates[slot]:
            opt_cal  = round(slot_cal * mult)          # each option has its own calorie value
            desc     = desc_fn(opt_cal)                # quantities scale from that option's kcal
            meals.append((name, desc, opt_cal))
        plan[slot] = meals

    tips = []
    if bmi > 27:
        deficit = round(target_cal * 0.15)
        tips.append(f"🔥 Your target is {target_cal} kcal/day — a ~{deficit} kcal deficit from maintenance. Avoid sugary drinks and fried foods.")
    elif bmi < 18.5:
        surplus = round(target_cal * 0.12)
        tips.append(f"📈 Your target is {target_cal} kcal/day — includes a ~{surplus} kcal surplus for healthy weight gain.")
    else:
        tips.append(f"⚖️ Your daily calorie target is {target_cal} kcal. Meal portions below are scaled to fit this.")

    if goal == "Muscle Gain":
        tips.append(f"💪 Target {round(1.6*weight)}–{round(2.0*weight)}g protein/day for your body weight ({weight} kg).")
        tips.append("🏋️ Consume a protein-rich meal within 30 min post-workout.")
    elif goal == "Weight Loss":
        tips.append("🥗 Prioritize high-volume, low-calorie foods (veggies, soups, salads) to stay full.")
        tips.append("🚰 Drink 2.5–3L water daily — hydration reduces false hunger signals.")
    else:
        tips.append("🕐 Eat every 3–4 hours to maintain stable blood sugar and energy.")

    if age > 40:
        tips.append("❤️ Prioritize calcium-rich foods (dairy, leafy greens) and limit sodium for heart health.")
    elif age < 20:
        tips.append("🌱 Ensure adequate calcium & iron intake for healthy growth and development.")

    return plan, tips

# -------------------------
# WEEKLY PLAN GENERATOR
# -------------------------
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

def generate_weekly_plan(goal, rda):
    """
    Builds a 7-day plan. Each day picks one option per slot, rotating through
    the 3 alternatives so meals don't repeat back-to-back.
    """
    fractions = MEAL_SLOT_FRACTIONS.get(goal, MEAL_SLOT_FRACTIONS["Maintain"])
    templates = MEAL_TEMPLATES.get(goal, MEAL_TEMPLATES["Maintain"])
    slots     = list(fractions.keys())

    # pre-build all option tuples per slot
    slot_options = {}
    for slot, frac in fractions.items():
        slot_cal = rda["calories"] * frac
        slot_options[slot] = [
            (name, desc_fn(round(slot_cal * mult)), round(slot_cal * mult))
            for mult, name, desc_fn in templates[slot]
        ]

    weekly = []
    for d_idx, day in enumerate(DAYS):
        day_meals = {}
        day_total = 0
        for slot in slots:
            options  = slot_options[slot]
            # rotate: Mon=opt0, Tue=opt1, Wed=opt2, Thu=opt0, ...
            pick_idx = d_idx % len(options)
            name, desc, kcal = options[pick_idx]
            day_meals[slot] = {"name": name, "desc": desc, "kcal": kcal}
            day_total += kcal
        weekly.append({"day": day, "meals": day_meals, "total_kcal": day_total})
    return weekly, slots


def export_pdf(weekly, slots, goal, weight, rda):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_fill_color(15, 17, 23)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(230, 230, 230)
    pdf.cell(0, 12, "Weekly Diet Plan", ln=True, align="C")
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(120, 140, 173)
    pdf.cell(0, 7, f"Goal: {goal}  |  Weight: {weight}kg  |  Daily Target: {rda['calories']} kcal", ln=True, align="C")
    pdf.ln(6)

    slot_colors = {
        "Breakfast": (59, 130, 246),
        "Lunch":     (16, 185, 129),
        "Dinner":    (245, 158, 11),
        "Snacks":    (139, 92, 246),
    }

    for day_data in weekly:
        day   = day_data["day"]
        meals = day_data["meals"]
        total = day_data["total_kcal"]

        # Day header bar
        pdf.set_fill_color(30, 37, 53)
        pdf.set_text_color(147, 197, 253)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 9, f"  {day}  —  {total} kcal", ln=True, fill=True)
        pdf.ln(2)

        for slot in slots:
            if slot not in meals:
                continue
            m = meals[slot]
            r, g, b = slot_colors.get(slot, (100, 100, 100))

            # Slot label
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(r, g, b)
            pdf.cell(28, 6, f"  {slot}", ln=False)

            # Meal name + kcal
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_text_color(220, 220, 220)
            pdf.cell(0, 6, f"{m['name']}  ({m['kcal']} kcal)", ln=True)

            # Description
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(160, 170, 190)
            pdf.set_x(28)
            pdf.multi_cell(0, 5, m["desc"])
            pdf.ln(1)

        pdf.ln(4)

    return bytes(pdf.output())


def export_excel(weekly, slots, goal, weight, rda):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Diet Plan"

    # Colour palette
    hdr_fill   = PatternFill("solid", fgColor="0F1117")
    day_fill   = PatternFill("solid", fgColor="1E2535")
    slot_fills = {
        "Breakfast": PatternFill("solid", fgColor="1D3B6E"),
        "Lunch":     PatternFill("solid", fgColor="0D3B2E"),
        "Dinner":    PatternFill("solid", fgColor="3B2A0A"),
        "Snacks":    PatternFill("solid", fgColor="2A1A4E"),
    }
    slot_fonts = {
        "Breakfast": Font(bold=True, color="93C5FD", size=10),
        "Lunch":     Font(bold=True, color="6EE7B7", size=10),
        "Dinner":    Font(bold=True, color="FCD34D", size=10),
        "Snacks":    Font(bold=True, color="C4B5FD", size=10),
    }
    thin = Side(style="thin", color="2E3A55")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row = 1
    # Title row
    ws.merge_cells(f"A{row}:F{row}")
    ws[f"A{row}"] = f"Weekly Diet Plan — {goal} | {weight}kg | {rda['calories']} kcal/day"
    ws[f"A{row}"].font      = Font(bold=True, color="E2E8F0", size=14)
    ws[f"A{row}"].fill      = hdr_fill
    ws[f"A{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 26
    row += 1

    # Column headers
    headers = ["Day", "Slot", "Meal Name", "Description", "Slot kcal", "Day Total"]
    col_widths = [14, 12, 30, 60, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font      = Font(bold=True, color="7B8CAD", size=10)
        cell.fill      = PatternFill("solid", fgColor="1A1F2E")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[row].height = 18
    row += 1

    for day_data in weekly:
        day   = day_data["day"]
        meals = day_data["meals"]
        total = day_data["total_kcal"]
        first_slot_row = row

        for si, slot in enumerate(slots):
            if slot not in meals:
                continue
            m = meals[slot]

            ws.cell(row=row, column=1, value=day if si == 0 else "").fill      = day_fill
            ws.cell(row=row, column=1).font      = Font(bold=True, color="E2E8F0", size=11)
            ws.cell(row=row, column=1).alignment = Alignment(vertical="center")
            ws.cell(row=row, column=1).border    = border

            ws.cell(row=row, column=2, value=slot).fill      = slot_fills.get(slot, day_fill)
            ws.cell(row=row, column=2).font      = slot_fonts.get(slot, Font(color="FFFFFF", size=10))
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=2).border    = border

            ws.cell(row=row, column=3, value=m["name"]).fill  = slot_fills.get(slot, day_fill)
            ws.cell(row=row, column=3).font      = Font(bold=True, color="E2E8F0", size=10)
            ws.cell(row=row, column=3).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row=row, column=3).border    = border

            ws.cell(row=row, column=4, value=m["desc"]).fill  = slot_fills.get(slot, day_fill)
            ws.cell(row=row, column=4).font      = Font(color="CBD5E1", size=9)
            ws.cell(row=row, column=4).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row=row, column=4).border    = border

            ws.cell(row=row, column=5, value=m["kcal"]).fill  = slot_fills.get(slot, day_fill)
            ws.cell(row=row, column=5).font      = Font(color="FBBF24", bold=True, size=10)
            ws.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=5).border    = border

            if si == 0:
                ws.cell(row=row, column=6, value=total).fill  = day_fill
                ws.cell(row=row, column=6).font      = Font(bold=True, color="6EE7B7", size=11)
                ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=6).border = border

            ws.row_dimensions[row].height = 30
            row += 1

        # Merge Day column and Total column for this day block
        if len(slots) > 1:
            last_slot_row = row - 1
            ws.merge_cells(f"A{first_slot_row}:A{last_slot_row}")
            ws.merge_cells(f"F{first_slot_row}:F{last_slot_row}")
            ws[f"A{first_slot_row}"].alignment = Alignment(horizontal="center", vertical="center")
            ws[f"F{first_slot_row}"].alignment = Alignment(horizontal="center", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# -------------------------
# CHARTS
# -------------------------
def plot_nutrient_rings(total, rda):
    carbs_p   = min(total["carbs"]   / max(rda["carbs"],   1), 1)
    protein_p = min(total["protein"] / max(rda["protein"], 1), 1)
    fat_p     = min(total["fat"]     / max(rda["fat"],     1), 1)
    cal_p     = min(total["calories"]/ max(rda["calories"],1), 1)

    progress = [carbs_p, protein_p, fat_p, cal_p]
    colors   = ["#3b82f6", "#10b981", "#f59e0b", "#ef4444"]
    labels   = ["Carbs", "Protein", "Fat", "Calories"]

    fig, ax = plt.subplots(figsize=(5, 5))
    fig.patch.set_facecolor("#0f1117")
    ax.set_facecolor("#0f1117")

    for i, (p, c) in enumerate(zip(progress, colors)):
        ax.pie([p, 1-p], radius=1 - i*0.2, startangle=90,
               counterclock=False, colors=[c, "#1e2535"],
               wedgeprops=dict(width=0.15, edgecolor="none"))

    ax.add_artist(plt.Circle((0, 0), 0.25, color="#0f1117"))
    patches = [mpatches.Patch(color=c, label=f"{l} {round(p*100)}%")
               for c, l, p in zip(colors, labels, progress)]
    ax.legend(handles=patches, loc="lower center", ncol=2,
              facecolor="#1e2535", edgecolor="#2a3550",
              labelcolor="#e2e8f0", fontsize=10,
              bbox_to_anchor=(0.5, -0.18))
    ax.set(aspect="equal")
    ax.axis("off")
    return fig

def plot_macro_bar(total, rda):
    categories = ["Calories", "Protein\n(g)", "Carbs\n(g)", "Fat\n(g)"]
    consumed   = [total["calories"], total["protein"], total["carbs"], total["fat"]]
    recommended= [rda["calories"],   rda["protein"],   rda["carbs"],  rda["fat"]]

    x    = np.arange(len(categories))
    w    = 0.35
    fig, ax = plt.subplots(figsize=(8, 4))
    fig.patch.set_facecolor("#0f1117")
    ax.set_facecolor("#1e2535")

    bars1 = ax.bar(x - w/2, consumed,    w, label="Consumed",    color="#3b82f6", alpha=0.9, zorder=3)
    bars2 = ax.bar(x + w/2, recommended, w, label="Recommended", color="#10b981", alpha=0.9, zorder=3)

    ax.bar_label(bars1, fmt="%.0f", color="#93c5fd", fontsize=9, padding=3)
    ax.bar_label(bars2, fmt="%.0f", color="#6ee7b7", fontsize=9, padding=3)

    ax.set_xticks(x)
    ax.set_xticklabels(categories, color="#cbd5e1", fontsize=11)
    ax.tick_params(axis="y", colors="#7b8cad")
    ax.spines[["top","right","left","bottom"]].set_visible(False)
    ax.yaxis.grid(True, color="#2a3550", linewidth=0.5, zorder=0)
    ax.set_axisbelow(True)
    ax.legend(facecolor="#1e2535", edgecolor="#2a3550", labelcolor="#e2e8f0", fontsize=10)
    fig.tight_layout()
    return fig

def plot_nutrient_breakdown(details):
    if not details: return None
    foods   = [d[0].title()[:12] for d in details]
    calvals = [d[1]["calories"] for d in details]
    colors  = ["#3b82f6","#10b981","#f59e0b","#ef4444","#8b5cf6","#ec4899"][:len(foods)]

    fig, ax = plt.subplots(figsize=(5, 5))
    fig.patch.set_facecolor("#0f1117")
    wedges, texts, autotexts = ax.pie(
        calvals, labels=foods, autopct="%1.0f%%",
        colors=colors[:len(foods)], startangle=140,
        wedgeprops=dict(edgecolor="#0f1117", linewidth=2),
        pctdistance=0.78
    )
    for t in texts:      t.set_color("#cbd5e1"); t.set_fontsize(10)
    for t in autotexts:  t.set_color("white");   t.set_fontsize(9); t.set_fontweight("bold")
    ax.set_facecolor("#0f1117")
    return fig

# -------------------------
# AUDIO
# -------------------------
def text_to_audio(text):
    tts  = gTTS(text)
    file = tempfile.NamedTemporaryFile(delete=False, suffix=".mp3")
    tts.save(file.name)
    return file.name

# ==========================================
# MAIN UPLOAD SECTION
# ==========================================
st.markdown("<div class='section-header'>📸 Upload Food Images</div>", unsafe_allow_html=True)
uploaded_files = st.file_uploader(
    "Drag & drop food images here",
    accept_multiple_files=True,
    type=["jpg", "jpeg", "png", "webp"],
    label_visibility="collapsed"
)

if uploaded_files:
    all_foods = []

    # ---- Image grid + detection ----
    st.markdown("<div class='section-header'>🔍 Detected Foods</div>", unsafe_allow_html=True)
    img_cols = st.columns(min(len(uploaded_files), 4))
    for i, file in enumerate(uploaded_files):
        img  = Image.open(file)
        res  = classifier(img)
        food = res[0]["label"].replace("_", " ")
        conf = round(res[0]["score"] * 100, 1)
        all_foods.append(food)
        with img_cols[i % 4]:
            st.image(img, use_container_width=True)
            st.markdown(f"""
            <div style='text-align:center; margin-top:4px;'>
                <span class='food-tag'>{food.title()}</span><br>
                <span style='color:#7b8cad; font-size:12px;'>Confidence: {conf}%</span>
            </div>""", unsafe_allow_html=True)

    # ---- Compute totals ----
    total, details = calculate_total(all_foods)
    bmi, bmi_status = calculate_bmi(weight, height)
    rda = get_rda(age, weight, height, gender, activity)

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # SECTION 1: KEY METRICS ROW
    # ==========================================
    st.markdown("<div class='section-header'>📊 Your Health Snapshot</div>", unsafe_allow_html=True)

    m1, m2, m3, m4, m5 = st.columns(5)
    bmi_col = "#6ee7b7" if bmi_status=="Normal" else ("#fcd34d" if bmi_status in ["Overweight","Underweight"] else "#fca5a5")
    h_m2    = height / 100
    min_w   = round(18.5*(h_m2**2), 1)
    max_w   = round(24.9*(h_m2**2), 1)
    bmr_val = round(calculate_bmr(weight, height, age, gender))

    with m1:
        st.markdown(f"<div class='metric-card'><div class='label'>BMI</div><div class='value' style='color:{bmi_col}'>{bmi}</div><div class='unit'>{bmi_status}</div></div>", unsafe_allow_html=True)
    with m2:
        st.markdown(f"<div class='metric-card'><div class='label'>Ideal Weight</div><div class='value'>{min_w}–{max_w}</div><div class='unit'>kg</div></div>", unsafe_allow_html=True)
    with m3:
        st.markdown(f"<div class='metric-card'><div class='label'>Daily Calories</div><div class='value'>{rda['calories']}</div><div class='unit'>kcal target</div></div>", unsafe_allow_html=True)
    with m4:
        st.markdown(f"<div class='metric-card'><div class='label'>BMR</div><div class='value'>{bmr_val}</div><div class='unit'>kcal/day</div></div>", unsafe_allow_html=True)
    with m5:
        cal_pct = round(total['calories'] / max(rda['calories'], 1) * 100)
        cal_color = "#6ee7b7" if 80 <= cal_pct <= 110 else ("#fcd34d" if cal_pct < 80 else "#fca5a5")
        st.markdown(f"<div class='metric-card'><div class='label'>Calories Consumed</div><div class='value' style='color:{cal_color}'>{round(total['calories'])}</div><div class='unit'>{cal_pct}% of target</div></div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # SECTION 2: CHARTS
    # ==========================================
    st.markdown("<div class='section-header'>📈 Nutrition Analysis</div>", unsafe_allow_html=True)

    ch1, ch2, ch3 = st.columns([1.2, 2, 1.2])

    with ch1:
        st.markdown("<div style='text-align:center; color:#7b8cad; font-size:13px; margin-bottom:8px;'>INTAKE vs RDA RINGS</div>", unsafe_allow_html=True)
        st.pyplot(plot_nutrient_rings(total, rda))

    with ch2:
        st.markdown("<div style='text-align:center; color:#7b8cad; font-size:13px; margin-bottom:8px;'>CONSUMED vs RECOMMENDED</div>", unsafe_allow_html=True)
        st.pyplot(plot_macro_bar(total, rda))

    with ch3:
        pie_fig = plot_nutrient_breakdown(details)
        if pie_fig:
            st.markdown("<div style='text-align:center; color:#7b8cad; font-size:13px; margin-bottom:8px;'>CALORIE SPLIT BY FOOD</div>", unsafe_allow_html=True)
            st.pyplot(pie_fig)

    st.markdown("<br>", unsafe_allow_html=True)

    # Per food breakdown
    st.markdown("<div class='section-header'>🍽️ Per Food Nutrients</div>", unsafe_allow_html=True)
    pf_cols = st.columns(2)
    for i, (f, d) in enumerate(details):
        with pf_cols[i % 2]:
            st.markdown(f"""
            <div class='nutrient-row'>
                <span style='color:#93c5fd; font-weight:600;'>{f.title()}</span>
                <span style='color:#7b8cad; font-size:13px;'>
                    🔥 {round(d['calories'])} kcal &nbsp;|&nbsp;
                    🥩 {round(d['protein'],1)}g protein &nbsp;|&nbsp;
                    🌾 {round(d['carbs'],1)}g carbs &nbsp;|&nbsp;
                    🫒 {round(d['fat'],1)}g fat
                </span>
            </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # SECTION 3: HEALTH ANALYSIS + GOAL
    # ==========================================
    st.markdown("<div class='section-header'>🧠 Health Analysis</div>", unsafe_allow_html=True)

    ha1, ha2 = st.columns([3, 2])

    with ha1:
        feedback = evaluate_food_health(all_foods, total, bmi, goal, age)
        for tip in feedback:
            icon = "✅" if "normal" in tip.lower() or "healthy" in tip.lower() or "good" in tip.lower() else "⚠️"
            st.markdown(f"<div class='tip-box'>{tip}</div>", unsafe_allow_html=True)

        deficiencies = analyze_deficiency(total, rda)
        st.markdown("<br>**Nutritional Gaps:**", unsafe_allow_html=False)
        for label, desc in deficiencies:
            color = "#6ee7b7" if "✅" in label else ("#fcd34d" if "⚠️" in label else "#fca5a5")
            st.markdown(f"""
            <div style='background:#1c2333; border:1px solid #2a3550; border-left:4px solid {color};
                        border-radius:8px; padding:10px 16px; margin:5px 0;'>
                <span style='color:{color}; font-weight:600;'>{label}</span>
                <span style='color:#cbd5e1; font-size:13px;'> — {desc}</span>
            </div>""", unsafe_allow_html=True)

    with ha2:
        goal_msg, goal_level = evaluate_goal(bmi, goal)
        badge_map = {"green": "badge-green", "yellow": "badge-yellow", "red": "badge-red"}
        st.markdown(f"""
        <div style='background:#1c2333; border:1px solid #2a3550; border-radius:12px; padding:20px;'>
            <div style='color:#7b8cad; font-size:12px; margin-bottom:8px;'>GOAL SUITABILITY</div>
            <div style='font-size:15px; color:#e2e8f0; margin-bottom:14px;'>{goal_msg}</div>
            <div style='color:#7b8cad; font-size:12px; margin-bottom:4px;'>CURRENT GOAL</div>
            <span class='{badge_map[goal_level]}'>{goal}</span>
            <div style='margin-top:16px; color:#7b8cad; font-size:12px;'>ACTIVITY LEVEL</div>
            <div style='color:#93c5fd; font-size:14px; margin-top:4px;'>{activity}</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # SECTION 4: MEAL RECOMMENDATIONS ⭐ NEW
    # ==========================================
    st.markdown("<div class='section-header'>🍱 Personalized Meal Plan</div>", unsafe_allow_html=True)
    meal_plan, meal_tips = get_meal_recommendations(goal, bmi, age, weight, rda)

    # Tips banner
    for tip in meal_tips:
        st.markdown(f"<div class='tip-box'>{tip}</div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    meal_tabs = st.tabs(["🌅 Breakfast", "☀️ Lunch", "🌙 Dinner", "🥜 Snacks"])
    meal_keys = ["Breakfast", "Lunch", "Dinner", "Snacks"]

    option_labels = [
        ("⭐ Recommended", "#3b82f6"),
        ("🔽 Lighter",     "#10b981"),
        ("🔽 Lightest",    "#8b5cf6"),
    ]

    for tab, key in zip(meal_tabs, meal_keys):
        with tab:
            options = meal_plan[key]
            cols = st.columns(len(options))
            for i, (col, (name, desc, kcal)) in enumerate(zip(cols, options)):
                badge_label, badge_color = option_labels[min(i, 2)]
                with col:
                    st.markdown(f"""
                    <div class='meal-card'>
                        <div style='margin-bottom:8px;'>
                            <span style='background:{badge_color}22; color:{badge_color};
                                border:1px solid {badge_color}55; border-radius:20px;
                                padding:2px 10px; font-size:11px; font-weight:600;'>
                                {badge_label}
                            </span>
                        </div>
                        <h4>{name}</h4>
                        <p>{desc}</p>
                        <div style='margin-top:10px; color:#fbbf24; font-size:13px; font-weight:600;'>
                            🔥 ~{kcal} kcal
                        </div>
                    </div>""", unsafe_allow_html=True)

    # Daily calorie budget bar — sum of all slot targets = rda calories
    st.markdown("<br>", unsafe_allow_html=True)
    fracs       = MEAL_SLOT_FRACTIONS.get(goal, MEAL_SLOT_FRACTIONS["Maintain"])
    day_total   = sum(round(rda["calories"] * f) for f in fracs.values())
    slot_labels = list(fracs.keys())
    slot_cals   = [round(rda["calories"] * f) for f in fracs.values()]
    used_pct    = min(day_total / max(rda["calories"], 1) * 100, 100)

    slot_html = "".join([
        f"<div style='text-align:center'>"
        f"<div style='color:#7b8cad; font-size:11px;'>{s}</div>"
        f"<div style='color:#e2e8f0; font-size:14px; font-weight:600;'>{c} kcal</div>"
        f"</div>"
        for s, c in zip(slot_labels, slot_cals)
    ])

    st.markdown(f"""
    <div style='background:#1c2333; border:1px solid #2a3550; border-radius:10px; padding:16px 20px;'>
        <div style='display:flex; justify-content:space-between; margin-bottom:10px; align-items:center;'>
            <span style='color:#e2e8f0; font-size:14px; font-weight:600;'>📅 Daily Calorie Plan — {weight}kg · {goal}</span>
            <span style='color:#fbbf24; font-size:15px; font-weight:700;'>{day_total} / {rda["calories"]} kcal</span>
        </div>
        <div style='display:flex; justify-content:space-around; margin-bottom:12px;'>
            {slot_html}
        </div>
        <div style='background:#2a3550; border-radius:20px; height:10px;'>
            <div style='background:linear-gradient(90deg,#3b82f6,#10b981); width:{used_pct:.1f}%;
                        height:10px; border-radius:20px;'></div>
        </div>
        <div style='color:#7b8cad; font-size:12px; margin-top:6px;'>
            Meal portions shown above are scaled to your {rda["calories"]} kcal/day target
        </div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # SECTION 5: WEEKLY DIET PLANNER
    # ==========================================
    st.markdown("<div class='section-header'>📅 Weekly Diet Planner (Mon–Sun)</div>", unsafe_allow_html=True)

    weekly_plan, plan_slots = generate_weekly_plan(goal, rda)

    slot_colors_ui = {
        "Breakfast": ("#3b82f6", "🌅"),
        "Lunch":     ("#10b981", "☀️"),
        "Dinner":    ("#f59e0b", "🌙"),
        "Snacks":    ("#8b5cf6", "🥜"),
    }

    # Calendar-style grid: 7 columns, one per day
    day_cols = st.columns(7)
    for col, day_data in zip(day_cols, weekly_plan):
        with col:
            day   = day_data["day"]
            meals = day_data["meals"]
            total = day_data["total_kcal"]

            st.markdown(f"""
            <div style='background:#1a1f2e; border:1px solid #2e3a55; border-radius:10px;
                        padding:10px 8px; min-height:320px;'>
                <div style='text-align:center; font-size:13px; font-weight:700;
                            color:#93c5fd; border-bottom:1px solid #2e3a55;
                            padding-bottom:6px; margin-bottom:8px;'>
                    {day[:3].upper()}
                    <div style='font-size:11px; color:#fbbf24; font-weight:600;'>{total} kcal</div>
                </div>
            """, unsafe_allow_html=True)

            for slot in plan_slots:
                if slot not in meals:
                    continue
                m = meals[slot]
                color, icon = slot_colors_ui.get(slot, ("#ffffff", "🍽️"))
                st.markdown(f"""
                <div style='background:#1c2840; border-left:3px solid {color};
                            border-radius:6px; padding:6px 8px; margin-bottom:6px;'>
                    <div style='font-size:10px; color:{color}; font-weight:700; margin-bottom:2px;'>
                        {icon} {slot.upper()}
                    </div>
                    <div style='font-size:11px; color:#e2e8f0; font-weight:600; line-height:1.3;'>
                        {m['name']}
                    </div>
                    <div style='font-size:10px; color:#fbbf24; margin-top:2px;'>
                        {m['kcal']} kcal
                    </div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Weekly summary bar chart
    st.markdown("<div style='color:#7b8cad; font-size:13px; margin-bottom:8px;'>📊 Weekly Calorie Overview</div>", unsafe_allow_html=True)
    fig_week, ax_week = plt.subplots(figsize=(10, 2.5))
    fig_week.patch.set_facecolor("#0f1117")
    ax_week.set_facecolor("#1e2535")
    day_names  = [d["day"][:3] for d in weekly_plan]
    day_totals = [d["total_kcal"] for d in weekly_plan]
    bar_colors = ["#3b82f6" if abs(t - rda["calories"]) < 100 else
                  ("#10b981" if t < rda["calories"] else "#f59e0b") for t in day_totals]
    bars = ax_week.bar(day_names, day_totals, color=bar_colors, alpha=0.9, zorder=3, width=0.55)
    ax_week.axhline(rda["calories"], color="#ef4444", linewidth=1.2, linestyle="--", zorder=4, label=f"Target {rda['calories']} kcal")
    ax_week.bar_label(bars, fmt="%d", color="#e2e8f0", fontsize=9, padding=3)
    ax_week.tick_params(colors="#cbd5e1", labelsize=10)
    ax_week.spines[["top","right","left","bottom"]].set_visible(False)
    ax_week.yaxis.grid(True, color="#2a3550", linewidth=0.5, zorder=0)
    ax_week.set_axisbelow(True)
    ax_week.legend(facecolor="#1e2535", edgecolor="#2a3550", labelcolor="#e2e8f0", fontsize=9)
    fig_week.tight_layout()
    st.pyplot(fig_week)

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Download buttons ----
    st.markdown("<div style='color:#e2e8f0; font-size:14px; font-weight:600; margin-bottom:10px;'>⬇️ Download Weekly Plan</div>", unsafe_allow_html=True)
    dl1, dl2 = st.columns(2)

    with dl1:
        pdf_bytes = export_pdf(weekly_plan, plan_slots, goal, weight, rda)
        st.download_button(
            label="📄 Download as PDF",
            data=pdf_bytes,
            file_name=f"weekly_diet_plan_{goal.lower().replace(' ','_')}.pdf",
            mime="application/pdf",
            use_container_width=True
        )

    with dl2:
        xlsx_bytes = export_excel(weekly_plan, plan_slots, goal, weight, rda)
        st.download_button(
            label="📊 Download as Excel",
            data=xlsx_bytes,
            file_name=f"weekly_diet_plan_{goal.lower().replace(' ','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ==========================================
    # SECTION 6: AUDIO
    # ==========================================
    st.markdown("<div class='section-header'>🔊 Audio Health Advice</div>", unsafe_allow_html=True)
    audio_text = f"Health analysis for your meal. {' '.join(feedback[:3])} "
    for label, desc in deficiencies[:2]:
        audio_text += f"{label}: {desc} "
    audio_text += f"Your BMI is {bmi}, classified as {bmi_status}. {goal_msg}"

    audio_path = text_to_audio(audio_text)
    st.audio(audio_path)
    if os.path.exists(audio_path):
        os.remove(audio_path)

else:
    # ---- Empty state ----
    st.markdown("""
    <div style='text-align:center; padding:60px 20px; color:#7b8cad;'>
        <div style='font-size:64px; margin-bottom:16px;'>📸</div>
        <div style='font-size:20px; font-weight:600; color:#e2e8f0; margin-bottom:8px;'>Upload Food Images to Begin</div>
        <div style='font-size:15px;'>Supports JPG, PNG, WebP · Multiple images allowed · AI identifies food & fetches USDA nutrition data</div>
        <br>
        <div style='display:flex; justify-content:center; gap:20px; flex-wrap:wrap; margin-top:10px;'>
            <span style='background:#1e2535; border:1px solid #2e3a55; border-radius:8px; padding:8px 16px; font-size:13px;'>🔍 AI Food Detection</span>
            <span style='background:#1e2535; border:1px solid #2e3a55; border-radius:8px; padding:8px 16px; font-size:13px;'>📊 USDA Nutrition Data</span>
            <span style='background:#1e2535; border:1px solid #2e3a55; border-radius:8px; padding:8px 16px; font-size:13px;'>🧠 BMI + BMR Analysis</span>
            <span style='background:#1e2535; border:1px solid #2e3a55; border-radius:8px; padding:8px 16px; font-size:13px;'>🍱 Personalized Meal Plan</span>
            <span style='background:#1e2535; border:1px solid #2e3a55; border-radius:8px; padding:8px 16px; font-size:13px;'>📅 Weekly Diet Planner + PDF/Excel</span>
            <span style='background:#1e2535; border:1px solid #2e3a55; border-radius:8px; padding:8px 16px; font-size:13px;'>🔊 Audio Advice</span>
        </div>
    </div>
    """, unsafe_allow_html=True)
